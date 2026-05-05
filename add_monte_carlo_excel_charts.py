"""Add comparison charts to the Monte Carlo Excel workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font


CHART_SPECS = (
    ("Penalized score", "eval_mean_logdet", "Mean log det incl. invalid penalty", "A1"),
    ("Valid-only information", "eval_mean_valid_logdet", "Mean log det on valid scenarios", "I1"),
    ("Scenario reliability", "eval_valid_fraction", "Valid fraction", "A18"),
    ("Illuminated/active sensors", "eval_mean_active_sensors", "Mean active sensors", "I18"),
    ("Below-noise sensors", "eval_mean_below_noise_sensors", "Mean below-noise sensors", "A35"),
)


def find_header(headers: list[str], name: str) -> int:
    try:
        return headers.index(name) + 1
    except ValueError as exc:
        raise ValueError(f"summary sheet does not contain column {name!r}") from exc


def rows_as_records(ws) -> list[dict[str, object]]:
    headers = [str(cell.value) for cell in ws[1]]
    records: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        records.append(dict(zip(headers, row, strict=True)))
    return records


def write_chart_data(wb, records: list[dict[str, object]]):
    if "chart_data" in wb.sheetnames:
        del wb["chart_data"]
    ws = wb.create_sheet("chart_data")
    headers = ["domain", "sensor_count"]
    for _, metric, _, _ in CHART_SPECS:
        headers.append(metric)
    ws.append(headers)

    # Keep separate domains and plot score vs sensor_count per domain.
    domain_order = ["single_face", "multi_face", "multi_face_per_face"]
    domain_rank = {name: idx for idx, name in enumerate(domain_order)}
    filtered = [r for r in records if str(r.get("optimizer_message", "")).startswith("heuristic:") is False]
    filtered.sort(key=lambda r: (domain_rank.get(str(r["domain"]), 999), int(r["sensor_count"])))
    for record in filtered:
        row: list[object] = [record["domain"], int(record["sensor_count"])]
        for _, metric, _, _ in CHART_SPECS:
            row.append(record.get(metric))
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.sheet_state = "hidden"
    return ws


def add_line_chart(
    data_ws,
    charts_ws,
    title: str,
    domain_name: str,
    value_column: int,
    category_column: int,
    anchor: str,
    max_row: int,
    y_axis_title: str,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.style = 13
    chart.height = 7.2
    chart.width = 14.5
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = "Sensor count (N total)"
    chart.legend.position = "b"
    # Filtered series per domain: build a temporary hidden table region on the charts sheet.
    # Simpler: assume chart_data has rows grouped by domain and contiguous.
    # We'll add one series per domain by scanning blocks.
    # The actual series assembly is done in add_charts().
    charts_ws.add_chart(chart, anchor)


def write_ranking_sheet(wb, records: list[dict[str, object]]) -> None:
    if "ranking" in wb.sheetnames:
        del wb["ranking"]
    ws = wb.create_sheet("ranking", 2)
    headers = [
        "rank",
        "domain",
        "sensor_count",
        "eval_valid_fraction",
        "eval_mean_valid_logdet",
        "eval_mean_logdet",
        "verdict",
    ]
    ws.append(headers)
    def as_float(value: object, default: float = float("-inf")) -> float:
        if value is None:
            return default
        try:
            return float(value)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return default

    ranked = sorted(
        records,
        key=lambda record: (
            as_float(record.get("eval_valid_fraction")),
            as_float(record.get("eval_mean_valid_logdet")),
            as_float(record.get("eval_mean_logdet")),
        ),
        reverse=True,
    )
    for index, record in enumerate(ranked, start=1):
        valid_fraction = as_float(record.get("eval_valid_fraction"), default=0.0)
        if valid_fraction >= 0.98:
            verdict = "good"
        elif valid_fraction >= 0.90:
            verdict = "usable / borderline"
        else:
            verdict = "bad reliability"
        ws.append(
            [
                index,
                record["domain"],
                record["sensor_count"],
                valid_fraction,
                as_float(record.get("eval_mean_valid_logdet"), default=float("nan")),
                as_float(record.get("eval_mean_logdet"), default=float("nan")),
                verdict,
            ]
        )
    for cell in ws[1]:
        cell.font = Font(bold=True)


def add_charts(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    if "summary" not in wb.sheetnames:
        raise ValueError("workbook must contain a summary sheet")
    if "charts" in wb.sheetnames:
        del wb["charts"]
    ws = wb["summary"]
    records = rows_as_records(ws)
    data_ws = write_chart_data(wb, records)
    write_ranking_sheet(wb, records)
    charts_ws = wb.create_sheet("charts", 1)
    charts_ws["A1"] = "Monte Carlo layout comparison charts: x-axis is sensor count; lines compare single-face vs multi-face."
    charts_ws["A1"].font = Font(bold=True, size=14)

    headers = [str(cell.value) for cell in data_ws[1]]
    domain_col = find_header(headers, "domain")
    count_col = find_header(headers, "sensor_count")

    domains = ["single_face", "multi_face", "multi_face_per_face"]
    max_row = data_ws.max_row

    for chart_index, (title, metric, y_axis_title, anchor) in enumerate(CHART_SPECS):
        metric_col = find_header(headers, metric)
        chart = LineChart()
        chart.title = title
        chart.style = 13
        chart.height = 7.2
        chart.width = 14.5
        chart.y_axis.title = y_axis_title
        chart.x_axis.title = "Sensor count (N total)"
        chart.legend.position = "b"

        # Build series per domain by collecting row indices.
        for domain in domains:
            rows = [r for r in range(2, max_row + 1) if str(data_ws.cell(row=r, column=domain_col).value) == domain]
            if not rows:
                continue
            # Create a contiguous helper column block on chart_data for this series.
            # We place them far to the right to avoid clobbering real data.
            base_col = len(headers) + 2 + chart_index * 6 + domains.index(domain) * 2
            data_ws.cell(row=1, column=base_col, value=f"{domain}_N")
            data_ws.cell(row=1, column=base_col + 1, value=f"{domain}_{metric}")
            for i, src_row in enumerate(rows, start=2):
                data_ws.cell(row=i, column=base_col, value=data_ws.cell(row=src_row, column=count_col).value)
                data_ws.cell(row=i, column=base_col + 1, value=data_ws.cell(row=src_row, column=metric_col).value)

            categories = Reference(data_ws, min_col=base_col, min_row=2, max_row=1 + len(rows))
            values = Reference(data_ws, min_col=base_col + 1, min_row=1, max_row=1 + len(rows))
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(categories)

        charts_ws.add_chart(chart, anchor)

    wb.save(workbook_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "workbook",
        type=Path,
        nargs="?",
        default=Path("monte_carlo_results/monte_carlo_face_trials.xlsx"),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    add_charts(args.workbook)
    print(f"added charts to {args.workbook}")


if __name__ == "__main__":
    main()
