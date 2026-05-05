"""Add per-trial sensor layout scatter charts to the Monte Carlo workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font


FACE_ORDER = ("+z_panel", "+x_panel", "-x_panel", "+y_panel", "-y_panel")


def find_sensors_header_row(ws) -> int:
    """Find the header row for the sensors table.

    We look for a row that contains all required headers: sensor, face, x_m, y_m.
    """
    required = {"sensor", "face", "x_m", "y_m"}
    for row in range(1, min(ws.max_row, 400) + 1):
        headers = {ws.cell(row=row, column=col).value for col in range(1, min(ws.max_column, 60) + 1)}
        if required.issubset(headers):
            return row
    raise ValueError(f"could not find sensors table header row in sheet {ws.title!r}")


def column_index(ws, header_row: int, name: str) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == name:
            return col
    raise ValueError(f"could not find column {name!r} in sheet {ws.title!r}")


def add_layout_chart_to_trial(ws) -> None:
    # Locate the sensors table written by run_face_monte_carlo.py (header includes "sensor", "face", "x_m", "y_m").
    header_row = find_sensors_header_row(ws)
    face_col = column_index(ws, header_row, "face")
    x_col = column_index(ws, header_row, "x_m")
    y_col = column_index(ws, header_row, "y_m")

    # Read rows until an empty sensor cell.
    data_rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=row, column=x_col).value in (None, "") or ws.cell(row=row, column=y_col).value in (None, ""):
            break
        data_rows.append(row)
    if not data_rows:
        return

    # Create helper columns to store per-face contiguous XY pairs for chart series.
    helper_start = ws.max_column + 2
    ws.cell(row=header_row, column=helper_start, value="layout_helper").font = Font(bold=True)

    chart = ScatterChart()
    chart.title = "Sensor layout (x,y) by face"
    chart.style = 2
    chart.height = 14
    chart.width = 24
    chart.x_axis.title = "x (m)"
    chart.y_axis.title = "y (m)"
    chart.legend.position = "r"

    max_points = 0
    for face_index, face_name in enumerate(FACE_ORDER):
        rows_for_face = [r for r in data_rows if ws.cell(row=r, column=face_col).value == face_name]
        if not rows_for_face:
            continue
        max_points = max(max_points, len(rows_for_face))
        base_col = helper_start + face_index * 2
        ws.cell(row=header_row, column=base_col, value=f"{face_name}_x")
        ws.cell(row=header_row, column=base_col + 1, value=f"{face_name}_y")
        for i, src_row in enumerate(rows_for_face, start=1):
            ws.cell(row=header_row + i, column=base_col, value=ws.cell(row=src_row, column=x_col).value)
            ws.cell(row=header_row + i, column=base_col + 1, value=ws.cell(row=src_row, column=y_col).value)

        xvalues = Reference(ws, min_col=base_col, min_row=header_row + 1, max_row=header_row + len(rows_for_face))
        yvalues = Reference(ws, min_col=base_col + 1, min_row=header_row + 1, max_row=header_row + len(rows_for_face))
        series = Series(yvalues, xvalues, title=face_name)
        series.marker.symbol = "circle"
        series.marker.size = 5
        chart.series.append(series)

    # Place the chart near the top; if it overlaps, Excel will still allow moving it.
    ws.add_chart(chart, "K2")


def add_layout_charts(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    for name in wb.sheetnames:
        if name in {"summary", "charts", "ranking", "chart_data"}:
            continue
        ws = wb[name]
        add_layout_chart_to_trial(ws)
    wb.save(workbook_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "workbook",
        type=Path,
        nargs="?",
        default=Path("monte_carlo_results/monte_carlo_face_trials.xlsx"),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    add_layout_charts(args.workbook)
    print(f"added layout charts to {args.workbook}")


if __name__ == "__main__":
    main()
